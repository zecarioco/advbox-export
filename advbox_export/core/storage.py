from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Espelha as 17 colunas do export do painel AdvBox + "Estado" e "Prazo"
# (extensões nossas pra classificar concluída/aberta × no prazo/atrasada
# a partir do cruzamento completed × date_deadline × hoje).
COLUNAS: list[tuple[str, str]] = [
    ("Data criada", "_data_criada"),
    ("Prioridade", "_prioridade"),
    ("Data agendada", "_data"),
    ("Hora", "_hora"),
    ("Término", "_termino_data"),
    ("Hora Término", "_termino_hora"),
    ("Prazo fatal", "_prazo_fatal"),
    ("Data Conclusão", "_data_conclusao"),
    ("Estado", "_estado"),
    ("Prazo", "_prazo_status"),
    ("Pontuação", "reward"),
    ("Compromisso", "task"),
    ("Local", "local"),
    ("Remetente", "_remetente"),
    ("Destinatário", "_destinatario"),
    ("Partes", "_partes"),
    ("Processo (CNJ)", "_process_number"),
    ("Protocolo", "_protocol_number"),
    ("Tipo de ação", "_tipo_acao"),
    ("Observações", "notes"),
]

# Estado: completou ou ainda está em aberto
ESTADO_CONCLUIDA = "Concluída"
ESTADO_EM_ABERTO = "Em aberto"

# Prazo: relação entre conclusão (ou hoje, se ainda aberta) e o deadline
PRAZO_NO_PRAZO = "No prazo"
PRAZO_ATRASADA = "Atrasada"
PRAZO_SEM_PRAZO = "Sem prazo"


def _calcular_estado(completed_raw: Any) -> str:
    if isinstance(completed_raw, str) and completed_raw:
        return ESTADO_CONCLUIDA
    return ESTADO_EM_ABERTO


def _calcular_prazo(
    completed_raw: Any, date_agendada_raw: Any, hoje_iso: str
) -> str:
    """Compara conclusão (ou hoje) com a data agendada pra classificar o prazo.

    Critério: a tarefa foi agendada pra acontecer numa data; se foi concluída
    depois dessa data, é atrasada — independente do `date_deadline`.

    - sem date agendada              → Sem prazo
    - concluída <= data agendada     → No prazo
    - concluída > data agendada      → Atrasada
    - em aberto, agendada no futuro  → No prazo
    - em aberto, agendada já passou  → Atrasada
    """
    if not isinstance(date_agendada_raw, str) or not date_agendada_raw:
        return PRAZO_SEM_PRAZO
    agendada_dia = date_agendada_raw[:10]
    if isinstance(completed_raw, str) and completed_raw:
        completed_dia = completed_raw[:10]
        return PRAZO_NO_PRAZO if completed_dia <= agendada_dia else PRAZO_ATRASADA
    return PRAZO_ATRASADA if agendada_dia < hoje_iso else PRAZO_NO_PRAZO

# Ocultas no XLSX, fora do CSV. Servem pra você inspecionar atividades
# atípicas sem precisar abrir o JSONL bruto.
COLUNAS_DEBUG_XLSX: list[tuple[str, str]] = [
    ("ID", "id"),
    ("Atividade (JSON bruto)", "_raw_atividade"),
    ("Lawsuit (JSON bruto)", "_raw_lawsuit"),
]


def expandir_atividade(
    atividade: dict[str, Any],
    *,
    range_inicio_iso: str | None = None,
    range_fim_iso: str | None = None,
    usuarios_permitidos: set[str] | None = None,
    hoje_iso: str | None = None,
) -> Iterable[dict[str, Any]]:
    """Emite uma linha por user designado da atividade.

    Pra cada user no array `users`: emite uma linha. Se o user concluiu,
    Data Conclusão/Término ficam preenchidos; se não, ficam vazios e a
    coluna Status mostra 'Em aberto' (ou 'Em aberto atrasada' se o prazo
    já passou). Isso permite o export funcionar como relatório completo
    do período, não apenas como relatório de produtividade concluída.

    Os parâmetros range_inicio_iso/range_fim_iso são aceitos por
    compatibilidade mas não filtram — quem filtra é a API (via
    completed_start/end ou deadline_start/end dependendo do modo da
    janela em /posts).
    """
    if hoje_iso is None:
        hoje_iso = date.today().isoformat()

    lawsuit = atividade.get("lawsuit") or {}
    customers = lawsuit.get("customers") or []
    users = atividade.get("users") or []

    date_raw = atividade.get("date")
    data, hora = _split_datetime(date_raw)
    deadline_raw = atividade.get("date_deadline")

    base = {
        "id": atividade.get("id"),
        "_data_criada": _formatar_data(atividade.get("created_at")),
        "_data": data,
        "_hora": hora,
        "_prazo_fatal": _formatar_data(deadline_raw),
        "reward": atividade.get("reward"),
        "task": atividade.get("task"),
        "local": atividade.get("local"),
        "_remetente": atividade.get("__author__") or "",
        "_partes": ", ".join(c.get("name", "") for c in customers if c.get("name")),
        "_process_number": lawsuit.get("process_number"),
        "_protocol_number": lawsuit.get("protocol_number"),
        "_tipo_acao": (atividade.get("__lawsuit_extra__") or {}).get("type") or "",
        "notes": atividade.get("notes"),
        "_raw_atividade": json.dumps(
            {k: v for k, v in atividade.items() if not k.startswith("__")},
            ensure_ascii=False,
        ),
        "_raw_lawsuit": json.dumps(lawsuit, ensure_ascii=False) if lawsuit else "",
    }

    for u in users:
        nome = u.get("name")
        if not nome:
            continue
        if usuarios_permitidos is not None and nome not in usuarios_permitidos:
            continue

        completed = u.get("completed") if isinstance(u.get("completed"), str) else None

        fim_data, fim_hora = ("", "")
        data_conclusao = ""
        if completed:
            fim_data, fim_hora = _split_datetime(completed)
            data_conclusao = (
                f"{fim_data} {fim_hora}" if fim_data and fim_hora else (fim_data or "")
            )

        linha = dict(base)
        linha["_prioridade"] = _prioridade_do_user(u)
        linha["_destinatario"] = nome
        linha["_termino_data"] = fim_data
        linha["_termino_hora"] = fim_hora
        linha["_data_conclusao"] = data_conclusao
        linha["_estado"] = _calcular_estado(completed)
        linha["_prazo_status"] = _calcular_prazo(completed, date_raw, hoje_iso)
        # Ordenação: concluídas pela data de conclusão; abertas pelo prazo
        # (ou pela date) — assim relatório agrupa "feitas em ordem" + "abertas
        # por prazo iminente".
        linha["_sort_key"] = completed or deadline_raw or date_raw or ""
        yield linha


def _prioridade_do_user(user: dict[str, Any]) -> str:
    """Mesmo critério do painel — flags do user específico daquela linha."""
    if user.get("urgent"):
        return "URGENTE"
    if user.get("important"):
        return "IMPORTANTE"
    return "NORMAL"


def _split_datetime(raw: Any) -> tuple[str, str]:
    """'2025-05-12 17:00:00' -> ('12/05/2025', '17:00'). Aceita também só data.

    '00:00' vira string vazia: o painel da AdvBox trata meia-noite como
    "sem hora marcada" (verificado em 215+ tarefas com date '...T00:00:00').
    """
    if not raw or not isinstance(raw, str):
        return "", ""
    raw = raw.strip()
    parte_data, _, parte_hora = raw.partition(" ")
    if not parte_hora:
        parte_data, _, parte_hora = raw.partition("T")
    data_fmt = _formatar_data(parte_data)
    hora_fmt = parte_hora[:5] if parte_hora else ""
    if hora_fmt == "00:00":
        hora_fmt = ""
    return data_fmt, hora_fmt


def _formatar_data(raw: Any) -> str:
    """'2025-05-12' ou '2025-05-12 17:00:00' -> '12/05/2025'."""
    if not raw or not isinstance(raw, str):
        return ""
    parte_data = raw[:10]
    if len(parte_data) == 10 and parte_data[4] == "-" and parte_data[7] == "-":
        return f"{parte_data[8:10]}/{parte_data[5:7]}/{parte_data[0:4]}"
    return raw


def gravar_jsonl_append(path: Path, atividades: Iterable[dict[str, Any]]) -> int:
    """Append atividades brutas (1 por linha) no JSONL. Retorna quantas gravou."""
    path.parent.mkdir(parents=True, exist_ok=True)
    count = 0
    with path.open("a", encoding="utf-8") as f:
        for a in atividades:
            f.write(json.dumps(a, ensure_ascii=False) + "\n")
            count += 1
    return count


def ler_jsonl(path: Path) -> Iterable[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as f:
        for linha in f:
            linha = linha.strip()
            if linha:
                yield json.loads(linha)


def _coletar_linhas_ordenadas(
    jsonl_path: Path,
    *,
    range_inicio_iso: str | None,
    range_fim_iso: str | None,
    usuarios_permitidos: set[str] | None = None,
) -> list[dict[str, Any]]:
    """Lê o JSONL, deduplica por atividade.id, expande, ordena ASC.

    Como o exporter faz 2 queries por janela (uma por `completed_start/end`,
    outra por `deadline_start/end`), uma mesma atividade pode aparecer 2
    vezes no JSONL. Dedup por `id` mantém só a primeira ocorrência — o
    array `users` da atividade é igual em ambas as respostas, então não
    perdemos informação.

    A API entrega atividades em ordem ~DESC por id e o painel da AdvBox
    apresenta o export ordenado ASC por Data Conclusão. Usamos o
    `_sort_key` ISO (completed → deadline → date) pra ordenação estável.
    """
    linhas: list[dict[str, Any]] = []
    ids_vistos: set[Any] = set()
    for atividade in ler_jsonl(jsonl_path):
        aid = atividade.get("id")
        if aid is not None and aid in ids_vistos:
            continue
        if aid is not None:
            ids_vistos.add(aid)
        linhas.extend(
            expandir_atividade(
                atividade,
                range_inicio_iso=range_inicio_iso,
                range_fim_iso=range_fim_iso,
                usuarios_permitidos=usuarios_permitidos,
            )
        )
    linhas.sort(key=lambda l: (l.get("_sort_key") or "", l.get("id") or 0))
    return linhas


def gerar_xlsx(
    jsonl_path: Path,
    xlsx_path: Path,
    *,
    range_inicio_iso: str | None = None,
    range_fim_iso: str | None = None,
    usuarios_permitidos: set[str] | None = None,
) -> int:
    """Lê o JSONL e escreve XLSX expandindo 1 linha por user-completado."""
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Atividades"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E5BBA")

    # Cores das colunas Estado/Prazo — tons claros pra não competir com o texto.
    estado_fills = {
        ESTADO_CONCLUIDA: PatternFill("solid", fgColor="D1FAE5"),  # verde claro
        ESTADO_EM_ABERTO: PatternFill("solid", fgColor="FEF3C7"),  # amarelo claro
    }
    prazo_fills = {
        PRAZO_NO_PRAZO: PatternFill("solid", fgColor="D1FAE5"),  # verde claro
        PRAZO_ATRASADA: PatternFill("solid", fgColor="FEE2E2"),  # vermelho claro
        PRAZO_SEM_PRAZO: PatternFill("solid", fgColor="F3F4F6"),  # cinza claro
    }

    colunas_xlsx = COLUNAS + COLUNAS_DEBUG_XLSX

    for col_idx, (header, _) in enumerate(colunas_xlsx, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    ws.freeze_panes = "A2"
    # Altura padrão idêntica à do export do painel (sheetFormatPr defaultRowHeight=14.4).
    ws.sheet_format.defaultRowHeight = 14.4

    # Painel da AdvBox aplica wrap_text=False em todas as células — sem isso,
    # Excel/LibreOffice auto-expandem linhas que contêm '\n' (comum em
    # Observações). Reaproveita o mesmo objeto pra não inflar o styles.xml.
    alinhamento_sem_wrap = Alignment(
        horizontal="general", vertical="bottom", wrap_text=False, shrink_to_fit=False
    )

    linhas = _coletar_linhas_ordenadas(
        jsonl_path,
        range_inicio_iso=range_inicio_iso,
        range_fim_iso=range_fim_iso,
        usuarios_permitidos=usuarios_permitidos,
    )

    row_idx = 2
    for linha in linhas:
        for col_idx, (_, key) in enumerate(colunas_xlsx, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=linha.get(key))
            cell.alignment = alinhamento_sem_wrap
            if key == "_estado":
                fill = estado_fills.get(linha.get("_estado"))
                if fill:
                    cell.fill = fill
            elif key == "_prazo_status":
                fill = prazo_fills.get(linha.get("_prazo_status"))
                if fill:
                    cell.fill = fill
        # Cinta dupla: defaultRowHeight + height por linha. Alguns leitores
        # respeitam um, outros respeitam o outro — setando ambos garantimos
        # que '\n' embutido em Observações nunca expanda a linha.
        ws.row_dimensions[row_idx].height = 14.4
        row_idx += 1

    larguras_visiveis = [
        14,  # Data criada
        12,  # Prioridade
        14,  # Data agendada
        10,  # Hora
        12,  # Término
        14,  # Hora Término
        14,  # Prazo fatal
        20,  # Data Conclusão
        12,  # Estado
        12,  # Prazo
        12,  # Pontuação
        36,  # Compromisso
        18,  # Local
        28,  # Remetente
        28,  # Destinatário
        36,  # Partes
        26,  # Processo (CNJ)
        18,  # Protocolo
        26,  # Tipo de ação
        48,  # Observações
    ]
    larguras_debug = [12, 80, 80]  # ID, Atividade bruta, Lawsuit bruto
    for i, w in enumerate(larguras_visiveis + larguras_debug, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # colunas de debug: largas (caso desoculte) e ocultas por padrão
    primeira_debug = len(COLUNAS) + 1
    for i in range(primeira_debug, primeira_debug + len(COLUNAS_DEBUG_XLSX)):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = 80
        ws.column_dimensions[col_letter].hidden = True

    # AutoFilter no header — dropdowns em cada coluna visível pro user filtrar
    # sem regerar o export. Cobre só até a última coluna visível (ignora debug).
    if row_idx > 2:
        ultima_col_visivel = get_column_letter(len(COLUNAS))
        ws.auto_filter.ref = f"A1:{ultima_col_visivel}{row_idx - 1}"

    # Aba "Resumo" na frente com contagens totais + breakdown por destinatário.
    _montar_aba_resumo(wb, linhas, estado_fills, prazo_fills)

    wb.save(xlsx_path)
    return row_idx - 2


def _montar_aba_resumo(
    wb: Workbook,
    linhas: list[dict[str, Any]],
    estado_fills: dict[str, PatternFill],
    prazo_fills: dict[str, PatternFill],
) -> None:
    """Cria uma aba 'Resumo' antes da 'Atividades' com contagens agregadas.

    Blocos:
      1. Total geral
      2. Breakdown por (Estado × Prazo)
      3. Tabela por destinatário
    """
    ws = wb.create_sheet("Resumo", 0)

    titulo_font = Font(bold=True, size=14, color="FFFFFF")
    titulo_fill = PatternFill("solid", fgColor="2E5BBA")
    header_font = Font(bold=True)
    section_font = Font(bold=True, size=12)

    total = len(linhas)
    contagem_geral: dict[tuple[str, str], int] = {}
    por_destinatario: dict[str, dict[tuple[str, str], int]] = {}
    for l in linhas:
        chave = (l.get("_estado") or "", l.get("_prazo_status") or "")
        contagem_geral[chave] = contagem_geral.get(chave, 0) + 1
        dest = l.get("_destinatario") or "(sem destinatário)"
        por_destinatario.setdefault(dest, {})
        por_destinatario[dest][chave] = por_destinatario[dest].get(chave, 0) + 1

    row = 1
    cell = ws.cell(row=row, column=1, value="Resumo do export")
    cell.font = titulo_font
    cell.fill = titulo_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 2

    ws.cell(row=row, column=1, value="Total de linhas").font = header_font
    ws.cell(row=row, column=2, value=total)
    row += 2

    # Combinações possíveis (Estado × Prazo)
    combinacoes = [
        (ESTADO_CONCLUIDA, PRAZO_NO_PRAZO),
        (ESTADO_CONCLUIDA, PRAZO_ATRASADA),
        (ESTADO_CONCLUIDA, PRAZO_SEM_PRAZO),
        (ESTADO_EM_ABERTO, PRAZO_NO_PRAZO),
        (ESTADO_EM_ABERTO, PRAZO_ATRASADA),
        (ESTADO_EM_ABERTO, PRAZO_SEM_PRAZO),
    ]

    ws.cell(row=row, column=1, value="Distribuição").font = section_font
    row += 1
    for header_txt in ("Estado", "Prazo", "Qtd", "%"):
        c = ws.cell(row=row, column=("Estado", "Prazo", "Qtd", "%").index(header_txt) + 1, value=header_txt)
        c.font = header_font
    row += 1
    for estado, prazo in combinacoes:
        qtd = contagem_geral.get((estado, prazo), 0)
        pct = (qtd / total * 100) if total else 0
        c_estado = ws.cell(row=row, column=1, value=estado)
        c_prazo = ws.cell(row=row, column=2, value=prazo)
        ws.cell(row=row, column=3, value=qtd)
        ws.cell(row=row, column=4, value=f"{pct:.1f}%")
        fill_e = estado_fills.get(estado)
        if fill_e:
            c_estado.fill = fill_e
        fill_p = prazo_fills.get(prazo)
        if fill_p:
            c_prazo.fill = fill_p
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Por destinatário").font = section_font
    row += 1

    headers_dest = [
        "Destinatário",
        "Total",
        "Concluída no prazo",
        "Concluída atrasada",
        "Concluída sem prazo",
        "Em aberto no prazo",
        "Em aberto atrasada",
        "Em aberto sem prazo",
    ]
    for col_idx, h in enumerate(headers_dest, start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = header_font
    row += 1

    for dest in sorted(por_destinatario.keys()):
        contagens = por_destinatario[dest]
        ws.cell(row=row, column=1, value=dest)
        total_dest = sum(contagens.values())
        ws.cell(row=row, column=2, value=total_dest)
        ws.cell(row=row, column=3, value=contagens.get((ESTADO_CONCLUIDA, PRAZO_NO_PRAZO), 0))
        ws.cell(row=row, column=4, value=contagens.get((ESTADO_CONCLUIDA, PRAZO_ATRASADA), 0))
        ws.cell(row=row, column=5, value=contagens.get((ESTADO_CONCLUIDA, PRAZO_SEM_PRAZO), 0))
        ws.cell(row=row, column=6, value=contagens.get((ESTADO_EM_ABERTO, PRAZO_NO_PRAZO), 0))
        ws.cell(row=row, column=7, value=contagens.get((ESTADO_EM_ABERTO, PRAZO_ATRASADA), 0))
        ws.cell(row=row, column=8, value=contagens.get((ESTADO_EM_ABERTO, PRAZO_SEM_PRAZO), 0))
        row += 1

    ws.freeze_panes = "A2"
    larguras_resumo = [32, 12, 22, 22, 22, 22, 22, 22]
    for i, w in enumerate(larguras_resumo, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


@dataclass(frozen=True)
class PeriodoArquivo:
    slug: str  # ex: "2026-06" ou "2024-01_2026-06"
    inicio: date
    fim: date


def slug_periodo(inicio: date, fim: date) -> str:
    """Slug curto pro nome do arquivo.

    Mês fechado: '2026-06'. Range cruzando meses: '2024-01_2026-06'.
    """
    if inicio.year == fim.year and inicio.month == fim.month:
        return f"{inicio:%Y-%m}"
    return f"{inicio:%Y-%m}_{fim:%Y-%m}"


def proximo_caminho_versionado(diretorio: Path, slug: str, sufixo: str) -> Path:
    """Devolve exports/{slug}_atividades.{ext} ou _v2, _v3 se já existir."""
    diretorio.mkdir(parents=True, exist_ok=True)
    base = f"{slug}_atividades"
    candidato = diretorio / f"{base}{sufixo}"
    if not candidato.exists():
        return candidato
    n = 2
    while True:
        candidato = diretorio / f"{base}_v{n}{sufixo}"
        if not candidato.exists():
            return candidato
        n += 1
